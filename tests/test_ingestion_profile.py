from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook

from autodata_agent.core.schemas import DataSourceType
from autodata_agent.services.datasets import DatasetStore


def test_csv_ingestion_builds_profile(tmp_path):
    store = DatasetStore(tmp_path, max_upload_bytes=5_000_000)
    content = b"date,region,sales,profit\n2026-01-01,North,100,20\n2026-01-02,South,150,-5\n"

    info = store.put_file("orders.csv", content)

    assert info.source_type == DataSourceType.CSV
    assert info.row_count == 2
    assert info.column_count == 4
    assert [column.name for column in info.profile.columns] == ["date", "region", "sales", "profit"]


def test_excel_and_json_ingestion(tmp_path):
    store = DatasetStore(tmp_path, max_upload_bytes=5_000_000)
    df = pd.DataFrame(
        [
            {"date": "2026-01-01", "segment": "Consumer", "sales": 100.0},
            {"date": "2026-01-02", "segment": "Corporate", "sales": 200.0},
        ]
    )

    excel_buffer = io.BytesIO()
    df.to_excel(excel_buffer, index=False)
    excel_info = store.put_file("orders.xlsx", excel_buffer.getvalue())
    json_info = store.put_file("orders.json", df.to_json(orient="records").encode("utf-8"))

    assert excel_info.source_type == DataSourceType.EXCEL
    assert json_info.source_type == DataSourceType.JSON
    assert excel_info.profile.row_count == json_info.profile.row_count == 2
    assert "sheet_name" in [column.name for column in excel_info.profile.columns]


def test_excel_ingestion_includes_all_workbook_sheets(tmp_path):
    store = DatasetStore(tmp_path, max_upload_bytes=5_000_000)
    schedule_df = pd.DataFrame([{"day": "Monday", "class": "Database"}])
    hld_df = pd.DataFrame(
        [
            {"Session Number": 1, "Session Title": "System Design 101"},
            {"Session Number": 2, "Session Title": "Load Balancing and Consistent Hashing"},
        ]
    )

    excel_buffer = io.BytesIO()
    with pd.ExcelWriter(excel_buffer) as writer:
        schedule_df.to_excel(writer, sheet_name="Weekly Schedule", index=False)
        hld_df.to_excel(writer, sheet_name="High Level Design", index=False)

    info = store.put_file("syllabus.xlsx", excel_buffer.getvalue())
    stored = store.get(info.dataset_id)

    assert info.row_count == 3
    assert "sheet_name" in stored.dataframe.columns
    assert set(stored.dataframe["sheet_name"]) == {"Weekly Schedule", "High Level Design"}
    assert "System Design 101" in stored.dataframe["Session Title"].dropna().tolist()


def test_excel_ingestion_normalizes_merged_timetable_sheet(tmp_path):
    store = DatasetStore(tmp_path, max_upload_bytes=5_000_000)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Weekly Schedule"
    sheet.merge_cells("C1:D1")
    sheet["C1"] = "Group A"
    sheet["C2"] = "Monday"
    sheet["D2"] = "Tuesday"
    sheet["A3"] = "9:00 - 9:15 AM"
    sheet["A4"] = "9:15 - 9:30 AM"
    sheet["A5"] = "9:30 - 9:45 AM"
    sheet.merge_cells("C3:C5")
    sheet["C3"] = "High Level Design [Instructor]"
    sheet["D3"] = "Lunch"
    sheet.merge_cells("D4:D5")
    sheet["D4"] = "Database Lab"

    buffer = io.BytesIO()
    workbook.save(buffer)

    info = store.put_file("schedule.xlsx", buffer.getvalue())
    stored = store.get(info.dataset_id)

    assert list(stored.dataframe.columns) == [
        "sheet_name",
        "Group",
        "Day",
        "Start Time",
        "End Time",
        "Activity",
    ]
    assert stored.dataframe.to_dict(orient="records") == [
        {
            "sheet_name": "Weekly Schedule",
            "Group": "Group A",
            "Day": "Monday",
            "Start Time": "9:00 AM",
            "End Time": "9:45 AM",
            "Activity": "High Level Design [Instructor]",
        },
        {
            "sheet_name": "Weekly Schedule",
            "Group": "Group A",
            "Day": "Tuesday",
            "Start Time": "9:00 AM",
            "End Time": "9:15 AM",
            "Activity": "Lunch",
        },
        {
            "sheet_name": "Weekly Schedule",
            "Group": "Group A",
            "Day": "Tuesday",
            "Start Time": "9:15 AM",
            "End Time": "9:45 AM",
            "Activity": "Database Lab",
        },
    ]


def test_dataset_store_rehydrates_from_disk(tmp_path):
    upload_dir = tmp_path / "uploads"
    dataset_dir = tmp_path / "datasets"
    store = DatasetStore(upload_dir, max_upload_bytes=5_000_000, dataset_dir=dataset_dir)
    info = store.put_file("orders.csv", b"category,sales\nA,100\nB,50\n")

    restarted_store = DatasetStore(upload_dir, max_upload_bytes=5_000_000, dataset_dir=dataset_dir)
    restored = restarted_store.get(info.dataset_id)

    assert restored.profile.row_count == 2
    assert list(restored.dataframe.columns) == ["category", "sales"]

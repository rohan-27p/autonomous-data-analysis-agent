from __future__ import annotations

import io
import json
import re
import uuid
from dataclasses import dataclass
from typing import Any
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import create_engine, text

from autodata_agent.core.errors import ValidationAppError
from autodata_agent.core.schemas import (
    DatasetInfo,
    DatasetPreview,
    DatasetProfile,
    DataSourceType,
    SQLIngestRequest,
)
from autodata_agent.services.profiling import build_profile


DAY_NAMES = {"monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"}
TIME_RANGE_RE = re.compile(
    r"\b\d{1,2}:\d{2}\s*(?:AM|PM)?\s*-\s*\d{1,2}:\d{2}\s*(?:AM|PM)?\b",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class StoredDataset:
    dataset_id: str
    dataframe: pd.DataFrame
    source_name: str
    source_type: DataSourceType
    profile: DatasetProfile


class DatasetStore:
    def __init__(
        self,
        upload_dir: Path,
        max_upload_bytes: int,
        dataset_dir: Path | None = None,
    ) -> None:
        self.upload_dir = upload_dir
        self.dataset_dir = dataset_dir or upload_dir.parent / "datasets"
        self.max_upload_bytes = max_upload_bytes
        self._datasets: dict[str, StoredDataset] = {}
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        self.dataset_dir.mkdir(parents=True, exist_ok=True)

    def put_file(self, filename: str, content: bytes) -> DatasetInfo:
        if not content:
            raise ValidationAppError("empty_upload", "Uploaded file is empty.")
        if len(content) > self.max_upload_bytes:
            raise ValidationAppError(
                "upload_too_large",
                "Uploaded file exceeds the configured size limit.",
                details={"max_upload_bytes": self.max_upload_bytes},
            )

        suffix = Path(filename).suffix.lower()
        if suffix == ".csv":
            source_type = DataSourceType.CSV
            df = pd.read_csv(io.BytesIO(content))
        elif suffix in {".xlsx", ".xls"}:
            source_type = DataSourceType.EXCEL
            df = self._read_excel_workbook(content)
        elif suffix == ".json":
            source_type = DataSourceType.JSON
            df = pd.read_json(io.BytesIO(content))
        else:
            raise ValidationAppError(
                "unsupported_file_type",
                "Supported file types are CSV, Excel, and JSON.",
                details={"filename": filename},
            )

        saved_name = f"{uuid.uuid4().hex}_{Path(filename).name}"
        self.upload_dir.mkdir(parents=True, exist_ok=True)
        (self.upload_dir / saved_name).write_bytes(content)
        return self._store(df, Path(filename).name, source_type)

    def _read_excel_workbook(self, content: bytes) -> pd.DataFrame:
        sheets = pd.read_excel(io.BytesIO(content), sheet_name=None)
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        frames: list[pd.DataFrame] = []
        for sheet_name, sheet_df in sheets.items():
            timetable_df = self._extract_timetable_sheet(workbook[sheet_name], sheet_name)
            if timetable_df is not None:
                frames.append(timetable_df)
                continue
            if sheet_df.empty:
                continue
            sheet_df = sheet_df.dropna(how="all").copy()
            if sheet_df.empty:
                continue
            metadata_column = "sheet_name"
            if metadata_column in sheet_df.columns:
                metadata_column = "__sheet_name"
            sheet_df.insert(0, metadata_column, sheet_name)
            frames.append(sheet_df)

        if not frames:
            raise ValidationAppError("empty_dataset", "Workbook contains no rows.")
        return pd.concat(frames, ignore_index=True, sort=False)

    def _extract_timetable_sheet(self, worksheet: Worksheet, sheet_name: str) -> pd.DataFrame | None:
        merged_ranges = list(worksheet.merged_cells.ranges)
        if len(merged_ranges) < 3:
            return None

        values = self._worksheet_values_with_merged_cells(worksheet)
        day_cells = [
            (row_idx, col_idx, str(value).strip())
            for row_idx, row in enumerate(values)
            for col_idx, value in enumerate(row)
            if self._is_day_name(value)
        ]
        time_cells = [
            (row_idx, col_idx, str(value).strip())
            for row_idx, row in enumerate(values)
            for col_idx, value in enumerate(row)
            if self._is_time_range(value)
        ]
        if len(day_cells) < 2 or len(time_cells) < 3:
            return None

        day_header_row = max(
            {row_idx for row_idx, _, _ in day_cells},
            key=lambda row_idx: sum(1 for candidate_row, _, _ in day_cells if candidate_row == row_idx),
        )
        day_columns = {
            col_idx: day
            for row_idx, col_idx, day in day_cells
            if row_idx == day_header_row
        }
        if len(day_columns) < 2:
            return None

        time_columns = self._time_columns_for_rows(time_cells, start_row=day_header_row + 1)
        if not time_columns:
            return None

        merged_lookup = {
            (cell_range.min_row - 1, cell_range.min_col - 1): (
                cell_range.max_row - cell_range.min_row + 1,
                cell_range.max_col - cell_range.min_col + 1,
            )
            for cell_range in merged_ranges
        }
        covered_by_merge = set()
        for (row_idx, col_idx), (row_span, col_span) in merged_lookup.items():
            for covered_row in range(row_idx, row_idx + row_span):
                for covered_col in range(col_idx, col_idx + col_span):
                    if (covered_row, covered_col) != (row_idx, col_idx):
                        covered_by_merge.add((covered_row, covered_col))

        records: list[dict[str, Any]] = []
        for col_idx, day in day_columns.items():
            group = self._header_above_column(values, day_header_row, col_idx)
            for row_idx in range(day_header_row + 1, len(values)):
                if (row_idx, col_idx) in covered_by_merge:
                    continue
                value = self._clean_cell_value(values[row_idx][col_idx])
                if not value or self._is_day_name(value) or self._is_time_range(value):
                    continue
                row_span = merged_lookup.get((row_idx, col_idx), (1, 1))[0]
                start_time, end_time = self._time_bounds_for_rows(values, time_columns, row_idx, row_span)
                records.append(
                    {
                        "sheet_name": sheet_name,
                        "Group": group,
                        "Day": day,
                        "Start Time": start_time,
                        "End Time": end_time,
                        "Activity": value,
                    }
                )

        if not records:
            return None
        return pd.DataFrame.from_records(records)

    @staticmethod
    def _worksheet_values_with_merged_cells(worksheet: Worksheet) -> list[list[Any]]:
        values = [
            [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
            for row_idx in range(1, worksheet.max_row + 1)
        ]
        for cell_range in worksheet.merged_cells.ranges:
            merged_value = worksheet.cell(cell_range.min_row, cell_range.min_col).value
            for row_idx in range(cell_range.min_row - 1, cell_range.max_row):
                for col_idx in range(cell_range.min_col - 1, cell_range.max_col):
                    values[row_idx][col_idx] = merged_value
        return values

    @staticmethod
    def _clean_cell_value(value: Any) -> str:
        if value is None:
            return ""
        return re.sub(r"\s+", " ", str(value)).strip()

    @classmethod
    def _is_day_name(cls, value: Any) -> bool:
        return cls._clean_cell_value(value).lower() in DAY_NAMES

    @classmethod
    def _is_time_range(cls, value: Any) -> bool:
        return bool(TIME_RANGE_RE.search(cls._clean_cell_value(value)))

    @staticmethod
    def _time_columns_for_rows(
        time_cells: list[tuple[int, int, str]],
        *,
        start_row: int,
    ) -> list[int]:
        counts: dict[int, int] = {}
        for row_idx, col_idx, _ in time_cells:
            if row_idx >= start_row:
                counts[col_idx] = counts.get(col_idx, 0) + 1
        if not counts:
            return []
        threshold = max(2, max(counts.values()) // 2)
        return [col_idx for col_idx, count in counts.items() if count >= threshold]

    @classmethod
    def _header_above_column(cls, values: list[list[Any]], day_header_row: int, col_idx: int) -> str | None:
        for row_idx in range(day_header_row - 1, -1, -1):
            value = cls._clean_cell_value(values[row_idx][col_idx])
            if value and not cls._is_day_name(value) and not cls._is_time_range(value):
                return value
        return None

    @classmethod
    def _time_bounds_for_rows(
        cls,
        values: list[list[Any]],
        time_columns: list[int],
        row_idx: int,
        row_span: int,
    ) -> tuple[str | None, str | None]:
        start_label = cls._time_label_for_row(values, time_columns, row_idx)
        end_label = cls._time_label_for_row(values, time_columns, row_idx + row_span - 1)
        return cls._split_time_range(start_label, "start"), cls._split_time_range(end_label, "end")

    @classmethod
    def _time_label_for_row(cls, values: list[list[Any]], time_columns: list[int], row_idx: int) -> str | None:
        if row_idx < 0 or row_idx >= len(values):
            return None
        for col_idx in time_columns:
            value = values[row_idx][col_idx] if col_idx < len(values[row_idx]) else None
            if cls._is_time_range(value):
                return cls._clean_cell_value(value)
        return None

    @staticmethod
    def _split_time_range(value: str | None, side: str) -> str | None:
        if not value:
            return None
        match = TIME_RANGE_RE.search(value)
        if not match:
            return None
        start, end = re.split(r"\s*-\s*", match.group(0), maxsplit=1)
        start = start.strip()
        end = end.strip()
        end_period = re.search(r"\b(AM|PM)\b", end, re.IGNORECASE)
        if end_period and not re.search(r"\b(AM|PM)\b", start, re.IGNORECASE):
            start = f"{start} {end_period.group(1).upper()}"
        return start if side == "start" else end

    def put_sql(self, request: SQLIngestRequest) -> DatasetInfo:
        cleaned = request.query.strip().lower()
        if not cleaned.startswith("select"):
            raise ValidationAppError(
                "unsafe_sql_query",
                "Only SELECT queries are allowed for SQL ingestion.",
            )
        try:
            engine = create_engine(request.connection_uri)
            with engine.connect() as connection:
                df = pd.read_sql_query(text(request.query), connection)
        except Exception as exc:  # noqa: BLE001 - returned as structured API error
            raise ValidationAppError(
                "sql_ingestion_failed",
                "SQL ingestion failed.",
                details={"reason": str(exc)},
            ) from exc
        return self._store(df, request.source_name, DataSourceType.SQL)

    def get(self, dataset_id: str) -> StoredDataset:
        try:
            return self._datasets[dataset_id]
        except KeyError as exc:
            rehydrated = self._rehydrate(dataset_id)
            if rehydrated is not None:
                self._datasets[dataset_id] = rehydrated
                return rehydrated
            raise ValidationAppError(
                "dataset_not_found",
                "Dataset was not found in backend storage.",
                status_code=404,
                details={"dataset_id": dataset_id},
            ) from exc

    def preview(self, dataset_id: str, limit: int = 20) -> DatasetPreview:
        stored = self.get(dataset_id)
        df = stored.dataframe.head(limit)
        rows = json.loads(
            df.where(pd.notnull(df), None).to_json(orient="records", date_format="iso")
        )
        return DatasetPreview(
            dataset_id=dataset_id,
            columns=[str(col) for col in df.columns],
            rows=rows,
            row_count=len(stored.dataframe),
            returned_rows=len(rows),
        )

    def _store(
        self,
        df: pd.DataFrame,
        source_name: str,
        source_type: DataSourceType,
    ) -> DatasetInfo:
        if df.empty:
            raise ValidationAppError("empty_dataset", "Dataset contains no rows.")
        df = df.copy()
        df.columns = [str(column).strip() for column in df.columns]
        if not all(df.columns):
            raise ValidationAppError("invalid_columns", "Dataset contains blank column names.")
        dataset_id = uuid.uuid4().hex
        profile = build_profile(
            df,
            dataset_id=dataset_id,
            source_name=source_name,
            source_type=source_type,
        )
        self._datasets[dataset_id] = StoredDataset(
            dataset_id,
            df,
            source_name,
            source_type,
            profile,
        )
        self._persist(dataset_id, df, profile)
        return DatasetInfo(
            dataset_id=dataset_id,
            source_name=source_name,
            source_type=source_type,
            row_count=profile.row_count,
            column_count=profile.column_count,
            profile=profile,
        )

    def _persist(self, dataset_id: str, df: pd.DataFrame, profile: DatasetProfile) -> None:
        dataset_path = self.dataset_dir / f"{dataset_id}.csv"
        profile_path = self.dataset_dir / f"{dataset_id}.profile.json"
        df.to_csv(dataset_path, index=False)
        profile_path.write_text(profile.model_dump_json(indent=2), encoding="utf-8")

    def _rehydrate(self, dataset_id: str) -> StoredDataset | None:
        dataset_path = self.dataset_dir / f"{dataset_id}.csv"
        profile_path = self.dataset_dir / f"{dataset_id}.profile.json"
        if not dataset_path.exists() or not profile_path.exists():
            return None
        df = pd.read_csv(dataset_path)
        profile = DatasetProfile.model_validate_json(profile_path.read_text(encoding="utf-8"))
        return StoredDataset(
            dataset_id=dataset_id,
            dataframe=df,
            source_name=profile.source_name,
            source_type=profile.source_type,
            profile=profile,
        )
